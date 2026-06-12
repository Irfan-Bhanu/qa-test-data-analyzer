import csv
import os

def read_csv_data(file_name):

    base_dir = os.path.dirname(
        os.path.dirname(os.path.abspath(__file__))
    )

    file_path = os.path.join(
        base_dir,
        "testdata",
        file_name
    )

    data=[]

    with open(file_path,'r') as csvfile:
        reader = csv.reader(csvfile)
        next(reader)

        for row in reader:
            data.append(tuple(row))

    return data

import json


def read_json_data(file_name):

    base_dir = os.path.dirname(
        os.path.dirname(os.path.abspath(__file__))
    )

    file_path = os.path.join(
        base_dir,
        "testdata",
        file_name
    )

    with open(file_path,'r') as file:

        data=json.load(file)

    return data

from openpyxl import load_workbook


def read_excel_data(file_name):

    base_dir = os.path.dirname(
        os.path.dirname(
            os.path.abspath(__file__)
        )
    )

    file_path = os.path.join(
        base_dir,
        "testdata",
        file_name
    )

    workbook = load_workbook(file_path)

    sheet = workbook.active
    print("Sheet:", sheet.title)
    print("Max rows:", sheet.max_row)
    print("Max columns:", sheet.max_column)

    data=[]

    for row in sheet.iter_rows(
            min_row=2,
            values_only=True):

        data.append(row)

        #print("Row:", row)  # Add this

        #data.append(row)


    #print("Final:", data)  # Add this

    return data